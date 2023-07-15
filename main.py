from time import sleep
import datetime
import platform
import pandas as pd
from modulo import dateVerification, abatimento, popup_Confirmacao_Exclusao, popupError, popup_Confirmacao_Backup
from modulo import formataLista
from openpyxl import load_workbook
from kivy.app import App
from kivy.uix.screenmanager import Screen, ScreenManager
from kivy.core.window import Window
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.popup import Popup
from kivy.uix.button import Button
from kivy.uix.label import Label


# Verifica se o usuário está usando Windows
if platform.system() == "Windows":
    sistema_windows = True
else:
    sistema_windows = False

# Variável para testar inserções de dados
teste = False


class MenuPrincipal(Screen):
    """
    Menu com as opções principais, NovosRegistros, LimparDados, ConsultaDeListas,
    Criar Backup e FecharPrograma.
    """
    pass


class NovosRegistros(Screen):
    """
    Opção para inserir novos dados, porém se faz necessário escolher as
    lista que deseja inserir os dados,
    poderá escolher entre: RD Marcas, Perfumaria e Dermo.
    """
    pass


class RegistrosRDMarcas(Screen):
    """
    Opção do menu principal após clicar na opção de registros (RDMarcas).
    """

    def pega_input_rdmarcas(self):
        """
        --> Função para pegar os dados inseridos na opção 'REGISTROS' -> 'RDMARCAS'.
        :return: Retorna os dados devidamente formatados.
        """
        try:
            data = self.ids.data_input.text
            data = dateVerification(data)
            metaDia = float(self.ids.meta_input.text.replace('-', '').strip())
            vendaDia = float(self.ids.venda_input.text.replace('-', '').strip())

            # Exibição no terminal
            pd.set_option('display.max_columns', None)  # Exibe todas as colunas
            pd.set_option('display.max_rows', None)  # Exibe todas as linhas
            pd.set_option('display.width', 1000)  # Largura máxima da exibição

            # Carrega o arquivo
            df_lista_RDMarcas = pd.read_excel('storage/listaRDMarcas.xlsx')
            calc_lista_RDMarcas = pd.read_excel('storage/lista_calc_RDMarcas.xlsx')

            novoCalc = {
                'Meta': f'{metaDia}',
                'Venda': f'{vendaDia}',
            }

            # Insere o dado na lista
            calc_lista_RDMarcas.loc[len(calc_lista_RDMarcas)] = novoCalc

            metaAC = calc_lista_RDMarcas['Meta'].astype(float).sum()
            vendaAC = calc_lista_RDMarcas['Venda'].astype(float).sum()

            if vendaAC < metaAC:
                sobras = (metaAC - vendaAC)
            elif metaAC < vendaAC:
                sobras = (vendaAC - metaAC)
            else:
                sobras = 0

            if vendaAC != 0 and metaAC != 0:
                porcentagem = (vendaAC / metaAC) * 100
            else:
                porcentagem = 'Error'

            # Análise alcance de metas
            devedor = abatimento(metaAC, vendaAC)

            # Cria uma linha de inserção de dados
            novoDado = {
                'Data': f'{data}',
                'Meta': f'{metaDia:.2f}',
                'Meta.AC': f'{metaAC:.2f}',
                'Venda': f'{vendaDia:.2f}',
                'Venda.AC': f'{vendaAC:.2f}',
                'Sobras': f'{sobras:.2f}',
                'P': f'{porcentagem:.2f}'
            }

            # Insere o dado na lista
            df_lista_RDMarcas.loc[len(df_lista_RDMarcas)] = novoDado
            print(df_lista_RDMarcas)
            print(f'Meta: {metaAC}')
            print(f'Venda: {vendaAC}')

            # Salva o arquivo com o novo dado
            df_lista_RDMarcas.to_excel('storage/listaRDMarcas.xlsx', index=False)
            calc_lista_RDMarcas.to_excel('storage/lista_calc_RDMarcas.xlsx', index=False)

            # Limpa os dados anteriormente informados (Somente teste = False)
            if not teste:
                self.ids.data_input.text = ""
                self.ids.meta_input.text = ""
                self.ids.venda_input.text = ""

                # Baseado na variável (devedor) o sistema passará a situação da meta/vendas no popup
                if devedor == '-':
                    situacao = "Metas não atingidas!"
                else:
                    situacao = "Metas atingitas!"

                # Popup de resumo
                content = BoxLayout(orientation='vertical', padding=10)
                label = Label(text=f'Resumo Acumulado (RD-Marcas)\n\n'
                                   f'Meta: R$ {metaAC:.2f}\nVendas: R$ {vendaAC:.2f}\n'
                                   f'Sobras: {devedor}R$ {sobras:.2f}\nSituação: {situacao}\n')
                close_button = Button(text='Fechar', size_hint=(None, None), size=(313, 50))

                content.add_widget(label)
                content.add_widget(close_button)

                popup = Popup(title='Dados armazenados com Sucesso!', content=content, size_hint=(None, None),
                              size=(360, 280))
                close_button.bind(on_release=popup.dismiss)
                popup.open()

        except ValueError as error:
            print(error)
            content = BoxLayout(orientation='vertical', padding=10)
            label = Label(text='Campos não preenchidos!')
            close_button = Button(text='Fechar', size_hint=(None, None), size=(100, 50))

            content.add_widget(label)
            content.add_widget(close_button)

            popup = Popup(title='Aviso', content=content, size_hint=(None, None), size=(375, 200))
            close_button.bind(on_release=popup.dismiss)
            popup.open()

        except Exception as error:
            print(error)
            popupError()


class RegistrosPerfumaria(Screen):
    """
    Opção do menu principal após clicar na opção de registros (Perfumaria).
    """

    def pega_input_perfumaria(self):
        """
        --> Função para pegar os dados inseridos na opção 'REGISTROS' -> 'PERFUMARIA'.
        :return: Retorna os dados devidamente formatados.
        """
        try:
            data = self.ids.data_input.text
            data = dateVerification(data)
            metaDia = float(self.ids.meta_input.text.replace('-', '').strip())
            vendaDia = float(self.ids.venda_input.text.replace('-', '').strip())

            # Exibição no terminal
            pd.set_option('display.max_columns', None)  # Exibe todas as colunas
            pd.set_option('display.max_rows', None)  # Exibe todas as linhas
            pd.set_option('display.width', 1000)  # Largura máxima da exibição

            # Carrega o arquivo
            df_lista_Perfumaria = pd.read_excel('storage/listaPerfumaria.xlsx')
            calc_lista_Perfumaria = pd.read_excel('storage/lista_calc_Perfumaria.xlsx')

            novoCalc = {
                'Meta': f'{metaDia}',
                'Venda': f'{vendaDia}',
            }

            # Insere o dado na lista
            calc_lista_Perfumaria.loc[len(calc_lista_Perfumaria)] = novoCalc

            metaAC = calc_lista_Perfumaria['Meta'].astype(float).sum()
            vendaAC = calc_lista_Perfumaria['Venda'].astype(float).sum()

            if vendaAC < metaAC:
                sobras = (metaAC - vendaAC)
            elif metaAC < vendaAC:
                sobras = (vendaAC - metaAC)
            else:
                sobras = 0

            if vendaAC != 0 and metaAC != 0:
                porcentagem = (vendaAC / metaAC) * 100
            else:
                porcentagem = 'Error'

            # Análise alcance de metas
            devedor = abatimento(metaAC, vendaAC)

            # Cria uma linha de inserção de dados
            novoDado = {
                'Data': f'{data}',
                'Meta': f'{metaDia:.2f}',
                'Meta.AC': f'{metaAC:.2f}',
                'Venda': f'{vendaDia:.2f}',
                'Venda.AC': f'{vendaAC:.2f}',
                'Sobras': f'{sobras:.2f}',
                'P': f'{porcentagem:.2f}'
            }

            # Insere o dado na lista
            df_lista_Perfumaria.loc[len(df_lista_Perfumaria)] = novoDado
            print(df_lista_Perfumaria)
            print(f'Meta: {metaAC}')
            print(f'Venda: {vendaAC}')

            # Salva o arquivo com o novo dado
            df_lista_Perfumaria.to_excel('storage/listaPerfumaria.xlsx', index=False)
            calc_lista_Perfumaria.to_excel('storage/lista_calc_Perfumaria.xlsx', index=False)

            # Limpa os dados anteriormente informados (Somente teste = False)
            if not teste:
                self.ids.data_input.text = ""
                self.ids.meta_input.text = ""
                self.ids.venda_input.text = ""

                # Baseado na variável (devedor) o sistema passará a situação da meta/vendas no popup
                if devedor == '-':
                    situacao = "Metas não atingidas!"
                else:
                    situacao = "Metas atingitas!"

                # Popup de resumo
                content = BoxLayout(orientation='vertical', padding=10)
                label = Label(text=f'Resumo Acumulado (PERFUMARIA)\n\n'
                                   f'Meta: R$ {metaAC:.2f}\nVendas: R$ {vendaAC:.2f}\n'
                                   f'Sobras: {devedor}R$ {sobras:.2f}\nSituação: {situacao}\n')
                close_button = Button(text='Fechar', size_hint=(None, None), size=(313, 50))

                content.add_widget(label)
                content.add_widget(close_button)

                popup = Popup(title='Dados armazenados com Sucesso!', content=content, size_hint=(None, None),
                              size=(360, 280))
                close_button.bind(on_release=popup.dismiss)
                popup.open()

        except ValueError as error:
            print(error)
            content = BoxLayout(orientation='vertical', padding=10)
            label = Label(text='Campos não preenchidos!')
            close_button = Button(text='Fechar', size_hint=(None, None), size=(100, 50))

            content.add_widget(label)
            content.add_widget(close_button)

            popup = Popup(title='Aviso', content=content, size_hint=(None, None), size=(375, 200))
            close_button.bind(on_release=popup.dismiss)
            popup.open()

        except Exception as error:
            print(error)
            popupError()


class RegistrosDermo(Screen):
    """
    Opção do menu principal após clicar na opção de registros (Dermo).
    """

    def pega_input_dermo(self):
        """
        --> Função para pegar os dados inseridos na opção 'REGISTROS' -> 'DERMO'.
        :return: Retorna os dados devidamente formatados.
        """
        try:
            data = self.ids.data_input.text
            data = dateVerification(data)
            metaDia = float(self.ids.meta_input.text.replace('-', '').strip())
            vendaDia = float(self.ids.venda_input.text.replace('-', '').strip())
            pecaDia = int(self.ids.peca_input.text.replace('-', '').strip())

            # Exibição no terminal
            pd.set_option('display.max_columns', None)  # Exibe todas as colunas
            pd.set_option('display.max_rows', None)  # Exibe todas as linhas
            pd.set_option('display.width', 1000)  # Largura máxima da exibição

            # Carrega o arquivo
            df_lista_Dermo = pd.read_excel('storage/listaDermo.xlsx')
            calc_lista_Dermo = pd.read_excel('storage/lista_calc_Dermo.xlsx')

            novoCalc = {
                'Meta': f'{metaDia}',
                'Venda': f'{vendaDia}',
                'Pecas': f'{pecaDia}',
            }

            # Insere o dado na lista
            calc_lista_Dermo.loc[len(calc_lista_Dermo)] = novoCalc

            metaAC = calc_lista_Dermo['Meta'].astype(float).sum()
            vendaAC = calc_lista_Dermo['Venda'].astype(float).sum()
            pecaAC = calc_lista_Dermo['Pecas'].astype(float).sum()

            if vendaAC < metaAC:
                sobras = (metaAC - vendaAC)
            elif metaAC < vendaAC:
                sobras = (vendaAC - metaAC)
            else:
                sobras = 0

            if vendaAC != 0 and metaAC != 0:
                porcentagem = (vendaAC / metaAC) * 100
            else:
                porcentagem = 'Error'

            # Análise alcance de metas
            devedor = abatimento(metaAC, vendaAC)

            # Cria uma linha de inserção de dados
            novoDado = {
                'Data': f'{data}',
                'Meta': f'{metaDia:.2f}',
                'Meta.AC': f'{metaAC:.2f}',
                'Venda': f'{vendaDia:.2f}',
                'Venda.AC': f'{vendaAC:.2f}',
                'Pecas.AC': f'{pecaAC}',
                'Sobras': f'{sobras:.2f}',
                'P': f'{porcentagem:.2f}'
            }

            # Insere o dado na lista
            df_lista_Dermo.loc[len(df_lista_Dermo)] = novoDado
            print(df_lista_Dermo)
            print(f'Meta: {metaAC}')
            print(f'Venda: {vendaAC}')
            print(f'Peças: {pecaAC}')

            # Salva o arquivo com o novo dado
            df_lista_Dermo.to_excel('storage/listaDermo.xlsx', index=False)
            calc_lista_Dermo.to_excel('storage/lista_calc_Dermo.xlsx', index=False)

            # Limpa os dados anteriormente informados (Somente teste = False)
            if not teste:
                self.ids.data_input.text = ""
                self.ids.meta_input.text = ""
                self.ids.venda_input.text = ""
                self.ids.peca_input.text = ""

                # Baseado na variável (devedor) o sistema passará a situação da meta/vendas no popup
                if devedor == '-':
                    situacao = "Metas não atingidas!"
                else:
                    situacao = "Metas atingitas!"

                # Popup de resumo
                content = BoxLayout(orientation='vertical', padding=10)
                label = Label(text=f'Resumo Acumulado (DERMO)\n\n'
                                   f'Meta: R$ {metaAC:.2f}\nVendas: R$ {vendaAC:.2f}\nPeças: {pecaAC} Un\n'
                                   f'Sobras: {devedor}R$ {sobras:.2f}\nSituação: {situacao}\n')
                close_button = Button(text='Fechar', size_hint=(None, None), size=(313, 50))

                content.add_widget(label)
                content.add_widget(close_button)

                popup = Popup(title='Dados armazenados com Sucesso!', content=content, size_hint=(None, None),
                              size=(360, 280))
                close_button.bind(on_release=popup.dismiss)
                popup.open()

        except ValueError as error:
            print(error)
            content = BoxLayout(orientation='vertical', padding=10)
            label = Label(text='Campos não preenchidos!')
            close_button = Button(text='Fechar', size_hint=(None, None), size=(100, 50))

            content.add_widget(label)
            content.add_widget(close_button)

            popup = Popup(title='Aviso', content=content, size_hint=(None, None), size=(375, 200))
            close_button.bind(on_release=popup.dismiss)
            popup.open()

        except Exception as error:
            print(error)
            popupError()


class LimparDados(Screen):
    """
    Opção para limpar dos dados, porém se faz necessário escolher as
    lista que deseja limpar os dados,
    poderá escolher entre: RD Marcas, Perfumaria, Dermo ou todas ao mesmo tempo.
    """

    def apagarLista_popup(self):
        """
        --> Função que mostra um Popup de confirmação antes de prosseguir com a exclusão das listas.
        """

        content = BoxLayout(orientation='vertical', padding=10)
        label = Label(text='Confirma a exclusão das Listas?')

        close_button = Button(text='Cancelar', size_hint=(None, None), size=(313, 50))
        confirm_button = Button(text='Confirmar', size_hint=(None, None), size=(313, 50))

        content.add_widget(label)
        content.add_widget(close_button)
        content.add_widget(confirm_button)

        popup = Popup(title='Aviso', content=content, size_hint=(None, None), size=(360, 280))

        close_button.bind(on_release=popup.dismiss)

        confirm_button.bind(on_release=lambda btn: self.apagarLista())
        confirm_button.bind(on_release=popup.dismiss)
        popup.open()

    def apagarLista(self):
        """
        --> Função que apaga todas as listas após confirmar no popup o procedimento.
        """

        try:
            # Exclusão RD MARCAS
            # Carrega o arquivo
            lista = 'storage/listaRDMarcas.xlsx'
            calculo = 'storage/lista_calc_RDMarcas.xlsx'
            bk_lista = load_workbook(lista)
            bk_calculo = load_workbook(calculo)

            # Pega a primeira planilha do arquivo de lista
            sheet_lista = bk_lista.active

            # Pega a primeira planilha do arquivo de cálculo
            sheet_calculo = bk_calculo.active

            # Exclui as linhas tirando a primeira (Nome das colunas/Key)
            sheet_lista.delete_rows(2, sheet_lista.max_row)
            sheet_calculo.delete_rows(2, sheet_calculo.max_row)

            # Salva as alterações
            bk_lista.save(lista)
            bk_calculo.save(calculo)

            # Passa as alterações para uma variável
            df_lista_RDMarcas = pd.read_excel(lista)
            calc_lista_RDMarcas = pd.read_excel(calculo)

            # Salva o arquivo com as alterações no DataFrame
            df_lista_RDMarcas.to_excel('storage/listaRDMarcas.xlsx', index=False)
            calc_lista_RDMarcas.to_excel('storage/lista_calc_RDMarcas.xlsx', index=False)

            # Exclusão PERFUMARIA
            # Carrega o arquivo
            lista = 'storage/listaPerfumaria.xlsx'
            calculo = 'storage/lista_calc_Perfumaria.xlsx'
            bk_lista = load_workbook(lista)
            bk_calculo = load_workbook(calculo)

            # Pega a primeira planilha do arquivo de lista
            sheet_lista = bk_lista.active

            # Pega a primeira planilha do arquivo de cálculo
            sheet_calculo = bk_calculo.active

            # Exclui as linhas tirando a primeira (Nome das colunas/Key)
            sheet_lista.delete_rows(2, sheet_lista.max_row)
            sheet_calculo.delete_rows(2, sheet_calculo.max_row)

            # Salva as alterações
            bk_lista.save(lista)
            bk_calculo.save(calculo)

            # Passa as alterações para uma variável
            df_lista_Perfumaria = pd.read_excel(lista)
            calc_lista_Perfumaria = pd.read_excel(calculo)

            # Salva o arquivo com as alterações no DataFrame
            df_lista_Perfumaria.to_excel('storage/listaPerfumaria.xlsx', index=False)
            calc_lista_Perfumaria.to_excel('storage/lista_calc_Perfumaria.xlsx', index=False)

            # Exclusão Dermo
            # Carrega o arquivo
            lista = 'storage/listaDermo.xlsx'
            calculo = 'storage/lista_calc_Dermo.xlsx'
            bk_lista = load_workbook(lista)
            bk_calculo = load_workbook(calculo)

            # Pega a primeira planilha do arquivo de lista
            sheet_lista = bk_lista.active

            # Pega a primeira planilha do arquivo de cálculo
            sheet_calculo = bk_calculo.active

            # Exclui as linhas tirando a primeira (Nome das colunas/Key)
            sheet_lista.delete_rows(2, sheet_lista.max_row)
            sheet_calculo.delete_rows(2, sheet_calculo.max_row)

            # Salva as alterações
            bk_lista.save(lista)
            bk_calculo.save(calculo)

            # Passa as alterações para uma variável
            df_lista_Dermo = pd.read_excel(lista)
            calc_lista_Dermo = pd.read_excel(calculo)

            # Salva o arquivo com as alterações no DataFrame
            df_lista_Dermo.to_excel('storage/listaDermo.xlsx', index=False)
            calc_lista_Dermo.to_excel('storage/lista_calc_Dermo.xlsx', index=False)

            # POPUP DE FINALIZAÇÃO
            sleep(0.3)
            popup_Confirmacao_Exclusao()

        except Exception as error:
            print(error)
            popupError()


class LimparRD(Screen):
    """
    """

    def __init__(self, **kwargs):
        super().__init__()
        self.index_value = None
        self.max_lines = None
        self.df_lista_RDMarcas = None
        self.calc_lista_RDMarcas = None
        self.tipo_busca = None

    def apagarLista_popup_RDMarcas(self):
        """
        --> Função que mostra um Popup de confirmação antes de prosseguir com a exclusão da lista.
        """

        content = BoxLayout(orientation='vertical', padding=10)
        label = Label(text='Confirma a exclusão da Lista RD Marcas?')

        close_button = Button(text='Cancelar', size_hint=(None, None), size=(313, 50))
        confirm_button = Button(text='Confirmar', size_hint=(None, None), size=(313, 50))

        content.add_widget(label)
        content.add_widget(close_button)
        content.add_widget(confirm_button)

        popup = Popup(title='Aviso', content=content, size_hint=(None, None), size=(360, 280))

        close_button.bind(on_release=popup.dismiss)

        confirm_button.bind(on_release=lambda btn: self.apagarLista_RDMarca())
        confirm_button.bind(on_release=popup.dismiss)
        popup.open()

    def apagarLista_RDMarca(self):
        """
        --> Função que apaga a lista RDMarcas após confirmar no popup o procedimento.
        """

        try:
            # Exclusão RD MARCAS
            # Carrega o arquivo
            lista = 'storage/listaRDMarcas.xlsx'
            calculo = 'storage/lista_calc_RDMarcas.xlsx'
            bk_lista = load_workbook(lista)
            bk_calculo = load_workbook(calculo)

            # Pega a primeira planilha do arquivo de lista
            sheet_lista = bk_lista.active

            # Pega a primeira planilha do arquivo de cálculo
            sheet_calculo = bk_calculo.active

            # Exclui as linhas tirando a primeira (Nome das colunas/Key)
            sheet_lista.delete_rows(2, sheet_lista.max_row)
            sheet_calculo.delete_rows(2, sheet_calculo.max_row)

            # Salva as alterações
            bk_lista.save(lista)
            bk_calculo.save(calculo)

            # Passa as alterações para uma variável
            df_lista_RDMarcas = pd.read_excel(lista)
            calc_lista_RDMarcas = pd.read_excel(calculo)

            # Salva o arquivo com as alterações no DataFrame
            df_lista_RDMarcas.to_excel('storage/listaRDMarcas.xlsx', index=False)
            calc_lista_RDMarcas.to_excel('storage/lista_calc_RDMarcas.xlsx', index=False)

            # POPUP DE FINALIZAÇÃO
            sleep(0.3)
            popup_Confirmacao_Exclusao()

        except Exception as error:
            print(error)
            popupError()

    def buscarPesquisa(self):
        try:
            # Carregar o arquivo (O arquivo de calculo e o arquivo da lista de visualização)
            global linha_filtrada
            self.calc_lista_RDMarcas = pd.read_excel('storage/lista_calc_RDMarcas.xlsx')
            self.df_lista_RDMarcas = pd.read_excel('storage/listaRDMarcas.xlsx')

            # Verifica a quantidade máxima de linhas dentro do arquivo
            self.max_lines = len(self.calc_lista_RDMarcas)  # Os dois aquivos importados tem a mesma quantidade de linha

            # Localiza a linha com base no input do usuário
            busca = self.ids.research_input.text

            # Limpa o resultado das buscas anteriores
            self.ids.resultado_linha.text = ''

            if len(str(busca)) > 4:
                linha_filtrada = self.df_lista_RDMarcas[self.df_lista_RDMarcas['Data'] == busca]
                self.ids.busca_resultado.text = f'Informação encontrada!'
                self.ids.resultado_linha.text = f'{linha_filtrada}'
                self.tipo_busca = 'data'

            elif int(busca.isnumeric()) and int(busca) - 2 <= self.max_lines - 1 and int(busca) >= 2:  # busca - 2
                busca = int(busca) - 2
                linha_filtrada = self.df_lista_RDMarcas[self.df_lista_RDMarcas.index == busca]
                self.ids.busca_resultado.text = f'Informação encontrada!'
                self.ids.resultado_linha.text = f'{linha_filtrada}'
                self.tipo_busca = 'index'

            else:
                self.ids.busca_resultado.text = 'Informação não Localizada!'
                self.ids.finalizar_alteracao.text = ''
                self.ids.resultado_linha.text = ''
                self.tipo_busca = None

            self.index_value = linha_filtrada.index[0]
            print(self.index_value)

            return self.calc_lista_RDMarcas, self.df_lista_RDMarcas, self.max_lines, self.index_value

        except Exception as error:
            print(f'Houve um erro - {error}')
            self.ids.busca_resultado.text = 'Informação não Localizada!'

    def executarAlteracao(self):
        try:
            # Define a quantidade de repetições iniciais
            qnt = 1

            # Cálculo dos dados
            data = self.ids.data_input.text
            metaDia = float(self.ids.meta_input.text.replace('-', '').strip())
            vendaDia = float(self.ids.venda_input.text.replace('-', '').strip())
            calc_lista_RDMarcas = self.calc_lista_RDMarcas
            for _ in calc_lista_RDMarcas.iterrows():
                if qnt == 1:
                    qnt = qnt + 1
                    # Insere os valores (MetaDia/VendaDia),
                    # logo em seguida é feito o cálculo já pegando o valor alterado
                    self.calc_lista_RDMarcas.loc[self.index_value] = [metaDia, vendaDia]

                    metaAC = self.calc_lista_RDMarcas.loc[:self.index_value, 'Meta'].astype(float).sum()
                    vendaAC = self.calc_lista_RDMarcas.loc[:self.index_value, 'Venda'].astype(float).sum()
                    if vendaAC < metaAC:
                        sobras = (metaAC - vendaAC)
                    elif metaAC < vendaAC:
                        sobras = (vendaAC - metaAC)
                    else:
                        sobras = 0

                    if vendaAC != 0 and metaAC != 0:
                        porcentagem = (vendaAC / metaAC) * 100
                    else:
                        porcentagem = 'Error'

                    # Input de dados
                    novoDado = {
                        'Data': f'{data}',
                        'Meta': f'{metaDia:.2f}',
                        'Meta.AC': f'{metaAC:.2f}',
                        'Venda': f'{vendaDia:.2f}',
                        'Venda.AC': f'{vendaAC:.2f}',
                        'Sobras': f'{sobras:.2f}',
                        'P': f'{porcentagem:.2f}'
                    }

                    # Modifica os valores da linha (MetaDia/VendaDia) | Modifica os dados da
                    # linha por completo com os devidos cálculos
                    self.df_lista_RDMarcas.loc[self.index_value] = novoDado

                    # Salva o arquivo
                    self.df_lista_RDMarcas.to_excel('storage/listaRDMarcas.xlsx', index=False)
                    self.calc_lista_RDMarcas.to_excel('storage/lista_calc_RDMarcas.xlsx', index=False)

                elif self.max_lines >= self.index_value and self.index_value < self.max_lines:
                    self.index_value = self.index_value + 1

                    metaDia = self.calc_lista_RDMarcas.at[self.index_value, 'Meta']
                    vendaDia = self.calc_lista_RDMarcas.at[self.index_value, 'Venda']
                    data_atualizada = self.df_lista_RDMarcas.at[self.index_value, 'Data']
                    self.calc_lista_RDMarcas.loc[self.index_value] = [metaDia, vendaDia]

                    metaAC = self.calc_lista_RDMarcas.loc[:self.index_value, 'Meta'].astype(float).sum()
                    vendaAC = self.calc_lista_RDMarcas.loc[:self.index_value, 'Venda'].astype(float).sum()

                    if vendaAC < metaAC:
                        sobras = (metaAC - vendaAC)
                    elif metaAC < vendaAC:
                        sobras = (vendaAC - metaAC)
                    else:
                        sobras = 0

                    if vendaAC != 0 and metaAC != 0:
                        porcentagem = (vendaAC / metaAC) * 100
                    else:
                        porcentagem = 'Error'

                    # Input de dados
                    novoDado = {
                        'Data': f'{data_atualizada}',
                        'Meta': f'{metaDia:.2f}',
                        'Meta.AC': f'{metaAC:.2f}',
                        'Venda': f'{vendaDia:.2f}',
                        'Venda.AC': f'{vendaAC:.2f}',
                        'Sobras': f'{sobras:.2f}',
                        'P': f'{porcentagem:.2f}'
                    }

                    self.df_lista_RDMarcas.loc[self.index_value] = novoDado

                    # Salva o arquivo
                    self.df_lista_RDMarcas.to_excel('storage/listaRDMarcas.xlsx', index=False)
                    self.calc_lista_RDMarcas.to_excel('storage/lista_calc_RDMarcas.xlsx', index=False)

                    # Texto do label de confimação após alterações
                    self.ids.finalizar_alteracao.text = 'Alterações realizadas'

                    # Limpa os dados inseridos e coloca a data de alteração no campo de pesquisa
                    if self.tipo_busca == 'data':
                        self.ids.research_input.text = f'{data}'

                    # Limpa os inputs preenchidos
                    self.ids.data_input.text = ''
                    self.ids.meta_input.text = ''
                    self.ids.venda_input.text = ''

        except Exception as error:
            print(f'Houve um erro - {error}')
            if self.tipo_busca == 'data':
                self.ids.finalizar_alteracao.text = 'Faça outra Busca\n  Para atualizar'
            elif self.tipo_busca == 'index':
                self.buscarPesquisa()
            else:
                self.ids.finalizar_alteracao.text = 'Faça outra Busca\n  Para atualizar'


class LimparPERFUMARIA(Screen):
    """
    """

    def __init__(self, **kw):
        super().__init__()
        self.tipo_busca = None
        self.df_lista_Perfumaria = None
        self.index_value = None
        self.calc_lista_Perfumaria = None
        self.max_lines = None

    def apagarLista_popup_Perfumaria(self):
        """
        --> Função que mostra um Popup de confirmação antes de prosseguir com a exclusão da lista.
        """

        content = BoxLayout(orientation='vertical', padding=10)
        label = Label(text='Confirma a exclusão da Lista Perfumaria?')

        close_button = Button(text='Cancelar', size_hint=(None, None), size=(313, 50))
        confirm_button = Button(text='Confirmar', size_hint=(None, None), size=(313, 50))

        content.add_widget(label)
        content.add_widget(close_button)
        content.add_widget(confirm_button)

        popup = Popup(title='Aviso', content=content, size_hint=(None, None), size=(360, 280))

        close_button.bind(on_release=popup.dismiss)

        confirm_button.bind(on_release=lambda btn: self.apagarLista_Perfumaria())
        confirm_button.bind(on_release=popup.dismiss)
        popup.open()

    def apagarLista_Perfumaria(self):
        """
        --> Função que apaga a lista Perfumaria após confirmar no popup o procedimento.
        """
        try:
            # Exclusão PERFUMARIA
            # Carrega o arquivo
            lista = 'storage/listaPerfumaria.xlsx'
            calculo = 'storage/lista_calc_Perfumaria.xlsx'
            bk_lista = load_workbook(lista)
            bk_calculo = load_workbook(calculo)

            # Pega a primeira planilha do arquivo de lista
            sheet_lista = bk_lista.active

            # Pega a primeira planilha do arquivo de cálculo
            sheet_calculo = bk_calculo.active

            # Exclui as linhas tirando a primeira (Nome das colunas/Key)
            sheet_lista.delete_rows(2, sheet_lista.max_row)
            sheet_calculo.delete_rows(2, sheet_calculo.max_row)

            # Salva as alterações
            bk_lista.save(lista)
            bk_calculo.save(calculo)

            # Passa as alterações para uma variável
            df_lista_Perfumaria = pd.read_excel(lista)
            calc_lista_Perfumaria = pd.read_excel(calculo)

            # Salva o arquivo com as alterações no DataFrame
            df_lista_Perfumaria.to_excel('storage/listaPerfumaria.xlsx', index=False)
            calc_lista_Perfumaria.to_excel('storage/lista_calc_Perfumaria.xlsx', index=False)

            # POPUP DE FINALIZAÇÃO
            sleep(0.3)
            popup_Confirmacao_Exclusao()

        except Exception as error:
            print(error)
            popupError()

    def buscarPesquisa(self):
        try:
            # Carregar o arquivo (O arquivo de calculo e o arquivo da lista de visualização)
            global linha_filtrada
            self.calc_lista_Perfumaria = pd.read_excel('storage/lista_calc_Perfumaria.xlsx')
            self.df_lista_Perfumaria = pd.read_excel('storage/listaPerfumaria.xlsx')

            # Verifica a quantidade máxima de linhas dentro do arquivo
            self.max_lines = len(
                self.calc_lista_Perfumaria)  # Os dois aquivos importados tem a mesma quantidade de linha

            # Localiza a linha com base no input do usuário
            busca = self.ids.research_input.text

            # Limpa o resultado das buscas anteriores
            self.ids.resultado_linha.text = ''

            if len(str(busca)) > 4:
                linha_filtrada = self.df_lista_Perfumaria[self.df_lista_Perfumaria['Data'] == busca]
                self.ids.busca_resultado.text = f'Informação encontrada!'
                self.ids.resultado_linha.text = f'{linha_filtrada}'
                self.tipo_busca = 'data'

            elif int(busca.isnumeric()) and int(busca) - 2 <= self.max_lines - 1 and int(busca) >= 2:  # busca - 2
                busca = int(busca) - 2
                linha_filtrada = self.df_lista_Perfumaria[self.df_lista_Perfumaria.index == busca]
                self.ids.busca_resultado.text = f'Informação encontrada!'
                self.ids.resultado_linha.text = f'{linha_filtrada}'
                self.tipo_busca = 'index'

            else:
                self.ids.busca_resultado.text = 'Informação não Localizada!'
                self.ids.finalizar_alteracao.text = ''
                self.ids.resultado_linha.text = ''
                self.tipo_busca = None

            self.index_value = linha_filtrada.index[0]
            print(self.index_value)

            return self.calc_lista_Perfumaria, self.df_lista_Perfumaria, self.max_lines, self.index_value

        except Exception as error:
            print(f'Houve um erro - {error}')
            self.ids.busca_resultado.text = 'Informação não Localizada!'

    def executarAlteracao(self):
        try:
            # Define a quantidade de repetições iniciais
            qnt = 1

            # Cálculo dos dados
            data = self.ids.data_input.text
            metaDia = float(self.ids.meta_input.text.replace('-', '').strip())
            vendaDia = float(self.ids.venda_input.text.replace('-', '').strip())
            calc_lista_Perfumaria = self.calc_lista_Perfumaria
            for _ in calc_lista_Perfumaria.iterrows():
                if qnt == 1:
                    qnt = qnt + 1
                    # Insere os valores (MetaDia/VendaDia),
                    # logo em seguida é feito o cálculo já pegando o valor alterado
                    self.calc_lista_Perfumaria.loc[self.index_value] = [metaDia, vendaDia]

                    metaAC = self.calc_lista_Perfumaria.loc[:self.index_value, 'Meta'].astype(float).sum()
                    vendaAC = self.calc_lista_Perfumaria.loc[:self.index_value, 'Venda'].astype(float).sum()
                    if vendaAC < metaAC:
                        sobras = (metaAC - vendaAC)
                    elif metaAC < vendaAC:
                        sobras = (vendaAC - metaAC)
                    else:
                        sobras = 0

                    if vendaAC != 0 and metaAC != 0:
                        porcentagem = (vendaAC / metaAC) * 100
                    else:
                        porcentagem = 'Error'

                    # Input de dados
                    novoDado = {
                        'Data': f'{data}',
                        'Meta': f'{metaDia:.2f}',
                        'Meta.AC': f'{metaAC:.2f}',
                        'Venda': f'{vendaDia:.2f}',
                        'Venda.AC': f'{vendaAC:.2f}',
                        'Sobras': f'{sobras:.2f}',
                        'P': f'{porcentagem:.2f}'
                    }

                    # Modifica os valores da linha (MetaDia/VendaDia) | Modifica os dados da
                    # linha por completo com os devidos cálculos
                    self.df_lista_Perfumaria.loc[self.index_value] = novoDado

                    # Salva o arquivo
                    self.df_lista_Perfumaria.to_excel('storage/listaPerfumaria.xlsx', index=False)
                    self.calc_lista_Perfumaria.to_excel('storage/lista_calc_Perfumaria.xlsx', index=False)

                elif self.max_lines >= self.index_value and self.index_value < self.max_lines:
                    self.index_value = self.index_value + 1

                    metaDia = self.calc_lista_Perfumaria.at[self.index_value, 'Meta']
                    vendaDia = self.calc_lista_Perfumaria.at[self.index_value, 'Venda']
                    data_atualizada = self.df_lista_Perfumaria.at[self.index_value, 'Data']
                    self.calc_lista_Perfumaria.loc[self.index_value] = [metaDia, vendaDia]

                    metaAC = self.calc_lista_Perfumaria.loc[:self.index_value, 'Meta'].astype(float).sum()
                    vendaAC = self.calc_lista_Perfumaria.loc[:self.index_value, 'Venda'].astype(float).sum()

                    if vendaAC < metaAC:
                        sobras = (metaAC - vendaAC)
                    elif metaAC < vendaAC:
                        sobras = (vendaAC - metaAC)
                    else:
                        sobras = 0

                    if vendaAC != 0 and metaAC != 0:
                        porcentagem = (vendaAC / metaAC) * 100
                    else:
                        porcentagem = 'Error'

                    # Input de dados
                    novoDado = {
                        'Data': f'{data_atualizada}',
                        'Meta': f'{metaDia:.2f}',
                        'Meta.AC': f'{metaAC:.2f}',
                        'Venda': f'{vendaDia:.2f}',
                        'Venda.AC': f'{vendaAC:.2f}',
                        'Sobras': f'{sobras:.2f}',
                        'P': f'{porcentagem:.2f}'
                    }

                    self.df_lista_Perfumaria.loc[self.index_value] = novoDado

                    # Salva o arquivo
                    self.df_lista_Perfumaria.to_excel('storage/listaPerfumaria.xlsx', index=False)
                    self.calc_lista_Perfumaria.to_excel('storage/lista_calc_Perfumaria.xlsx', index=False)

                    # Texto do label de confimação após alterações
                    self.ids.finalizar_alteracao.text = 'Alterações realizadas'

                    # Limpa os dados inseridos e coloca a data de alteração no campo de pesquisa
                    if self.tipo_busca == 'data':
                        self.ids.research_input.text = f'{data}'

                    # Limpa os inputs preenchidos
                    self.ids.data_input.text = ''
                    self.ids.meta_input.text = ''
                    self.ids.venda_input.text = ''

        except Exception as error:
            print(f'Houve um erro - {error}')
            if self.tipo_busca == 'data':
                self.ids.finalizar_alteracao.text = 'Faça outra Busca\n  Para atualizar'
            elif self.tipo_busca == 'index':
                self.buscarPesquisa()
            else:
                self.ids.finalizar_alteracao.text = 'Faça outra Busca\n  Para atualizar'


class LimparDERMO(Screen):
    """
    """

    def __init__(self, **kw):
        super().__init__()
        self.tipo_busca = None
        self.index_value = None
        self.max_lines = None
        self.df_lista_Dermo = None
        self.calc_lista_Dermo = None

    def apagarLista_popup_Dermo(self):
        """
        --> Função que mostra um Popup de confirmação antes de prosseguir com a exclusão das listas.
        """

        content = BoxLayout(orientation='vertical', padding=10)
        label = Label(text='Confirma a exclusão da Lista Dermo?')

        close_button = Button(text='Cancelar', size_hint=(None, None), size=(313, 50))
        confirm_button = Button(text='Confirmar', size_hint=(None, None), size=(313, 50))

        content.add_widget(label)
        content.add_widget(close_button)
        content.add_widget(confirm_button)

        popup = Popup(title='Aviso', content=content, size_hint=(None, None), size=(360, 280))

        close_button.bind(on_release=popup.dismiss)

        confirm_button.bind(on_release=lambda btn: self.apagarLista_Dermo())
        confirm_button.bind(on_release=popup.dismiss)
        popup.open()

    def apagarLista_Dermo(self):
        """
        --> Função que apaga a lista Dermo após confirmar no popup o procedimento.
        """

        try:
            # Exclusão Dermo
            # Carrega o arquivo
            lista = 'storage/listaDermo.xlsx'
            calculo = 'storage/lista_calc_Dermo.xlsx'
            bk_lista = load_workbook(lista)
            bk_calculo = load_workbook(calculo)

            # Pega a primeira planilha do arquivo de lista
            sheet_lista = bk_lista.active

            # Pega a primeira planilha do arquivo de cálculo
            sheet_calculo = bk_calculo.active

            # Exclui as linhas tirando a primeira (Nome das colunas/Key)
            sheet_lista.delete_rows(2, sheet_lista.max_row)
            sheet_calculo.delete_rows(2, sheet_calculo.max_row)

            # Salva as alterações
            bk_lista.save(lista)
            bk_calculo.save(calculo)

            # Passa as alterações para uma variável
            df_lista_Dermo = pd.read_excel(lista)
            calc_lista_Dermo = pd.read_excel(calculo)

            # Salva o arquivo com as alterações no DataFrame
            df_lista_Dermo.to_excel('storage/listaDermo.xlsx', index=False)
            calc_lista_Dermo.to_excel('storage/lista_calc_Dermo.xlsx', index=False)

            # POPUP DE FINALIZAÇÃO
            sleep(0.3)
            popup_Confirmacao_Exclusao()

        except Exception as error:
            print(error)
            popupError()

    def buscarPesquisa(self):
        try:
            # Carregar o arquivo (O arquivo de calculo e o arquivo da lista de visualização)
            global linha_filtrada
            self.calc_lista_Dermo = pd.read_excel('storage/lista_calc_Dermo.xlsx')
            self.df_lista_Dermo = pd.read_excel('storage/listaDermo.xlsx')

            # Verifica a quantidade máxima de linhas dentro do arquivo
            self.max_lines = len(
                self.calc_lista_Dermo)  # Os dois aquivos importados tem a mesma quantidade de linha

            # Localiza a linha com base no input do usuário
            busca = self.ids.research_input.text

            # Limpa o resultado das buscas anteriores
            self.ids.resultado_linha.text = ''

            if len(str(busca)) > 4:
                linha_filtrada = self.df_lista_Dermo[self.df_lista_Dermo['Data'] == busca]
                self.ids.busca_resultado.text = f'Informação encontrada!'
                self.ids.resultado_linha.text = f'{linha_filtrada}'
                self.tipo_busca = 'data'

            elif int(busca.isnumeric()) and int(busca) - 2 <= self.max_lines - 1 and int(busca) >= 2:  # busca - 2
                busca = int(busca) - 2
                linha_filtrada = self.df_lista_Dermo[self.df_lista_Dermo.index == busca]
                self.ids.busca_resultado.text = f'Informação encontrada!'
                self.ids.resultado_linha.text = f'{linha_filtrada}'
                self.tipo_busca = 'index'

            else:
                self.ids.busca_resultado.text = 'Informação não Localizada!'
                self.ids.finalizar_alteracao.text = ''
                self.ids.resultado_linha.text = ''
                self.tipo_busca = None

            self.index_value = linha_filtrada.index[0]
            print(self.index_value)

            return self.calc_lista_Dermo, self.df_lista_Dermo, self.max_lines, self.index_value

        except Exception as error:
            print(f'Houve um erro - {error}')
            self.ids.busca_resultado.text = 'Informação não Localizada!'

    def executarAlteracao(self):
        try:
            # Define a quantidade de repetições iniciais
            qnt = 1

            # Cálculo dos dados
            data = self.ids.data_input.text
            metaDia = float(self.ids.meta_input.text.replace('-', '').strip())
            vendaDia = float(self.ids.venda_input.text.replace('-', '').strip())
            pecaDia = int(self.ids.peca_input.text.replace('-', '').strip())
            calc_lista_Dermo = self.calc_lista_Dermo
            for _ in calc_lista_Dermo.iterrows():
                if qnt == 1:
                    qnt = qnt + 1
                    # Insere os valores (MetaDia/VendaDia),
                    # logo em seguida é feito o cálculo já pegando o valor alterado
                    self.calc_lista_Dermo.loc[self.index_value] = [metaDia, vendaDia, pecaDia]

                    metaAC = self.calc_lista_Dermo.loc[:self.index_value, 'Meta'].astype(float).sum()
                    vendaAC = self.calc_lista_Dermo.loc[:self.index_value, 'Venda'].astype(float).sum()
                    pecaAC = self.calc_lista_Dermo.loc[:self.index_value, 'Pecas'].astype(int).sum()
                    if vendaAC < metaAC:
                        sobras = (metaAC - vendaAC)
                    elif metaAC < vendaAC:
                        sobras = (vendaAC - metaAC)
                    else:
                        sobras = 0

                    if vendaAC != 0 and metaAC != 0:
                        porcentagem = (vendaAC / metaAC) * 100
                    else:
                        porcentagem = 'Error'

                    # Input de dados
                    novoDado = {
                        'Data': f'{data}',
                        'Meta': f'{metaDia:.2f}',
                        'Meta.AC': f'{metaAC:.2f}',
                        'Venda': f'{vendaDia:.2f}',
                        'Venda.AC': f'{vendaAC:.2f}',
                        'Pecas.AC': f'{pecaAC}',
                        'Sobras': f'{sobras:.2f}',
                        'P': f'{porcentagem:.2f}'
                    }

                    # Modifica os valores da linha (MetaDia/VendaDia) | Modifica os dados da
                    # linha por completo com os devidos cálculos
                    self.df_lista_Dermo.loc[self.index_value] = novoDado

                    # Salva o arquivo
                    self.df_lista_Dermo.to_excel('storage/listaDermo.xlsx', index=False)
                    self.calc_lista_Dermo.to_excel('storage/lista_calc_Dermo.xlsx', index=False)

                elif self.max_lines >= self.index_value and self.index_value < self.max_lines:
                    self.index_value = self.index_value + 1

                    metaDia = self.calc_lista_Dermo.at[self.index_value, 'Meta']
                    vendaDia = self.calc_lista_Dermo.at[self.index_value, 'Venda']
                    pecaDia = self.calc_lista_Dermo.at[self.index_value, 'Pecas']

                    data_atualizada = self.df_lista_Dermo.at[self.index_value, 'Data']
                    self.calc_lista_Dermo.loc[self.index_value] = [metaDia, vendaDia, pecaDia]

                    metaAC = self.calc_lista_Dermo.loc[:self.index_value, 'Meta'].astype(float).sum()
                    vendaAC = self.calc_lista_Dermo.loc[:self.index_value, 'Venda'].astype(float).sum()
                    pecaAC = self.calc_lista_Dermo.loc[:self.index_value, 'Pecas'].astype(int).sum()

                    if vendaAC < metaAC:
                        sobras = (metaAC - vendaAC)
                    elif metaAC < vendaAC:
                        sobras = (vendaAC - metaAC)
                    else:
                        sobras = 0

                    if vendaAC != 0 and metaAC != 0:
                        porcentagem = (vendaAC / metaAC) * 100
                    else:
                        porcentagem = 'Error'

                    # Input de dados
                    novoDado = {
                        'Data': f'{data_atualizada}',
                        'Meta': f'{metaDia:.2f}',
                        'Meta.AC': f'{metaAC:.2f}',
                        'Venda': f'{vendaDia:.2f}',
                        'Venda.AC': f'{vendaAC:.2f}',
                        'Pecas.AC': f'{pecaAC}',
                        'Sobras': f'{sobras:.2f}',
                        'P': f'{porcentagem:.2f}'
                    }

                    self.df_lista_Dermo.loc[self.index_value] = novoDado

                    # Salva o arquivo
                    self.df_lista_Dermo.to_excel('storage/listaDermo.xlsx', index=False)
                    self.calc_lista_Dermo.to_excel('storage/lista_calc_Dermo.xlsx', index=False)

                    # Texto do label de confimação após alterações
                    self.ids.finalizar_alteracao.text = 'Alterações realizadas'

                    # Limpa os dados inseridos e coloca a data de alteração no campo de pesquisa
                    if self.tipo_busca == 'data':
                        self.ids.research_input.text = f'{data}'

                    # Limpa os inputs preenchidos
                    self.ids.data_input.text = ''
                    self.ids.meta_input.text = ''
                    self.ids.venda_input.text = ''
                    self.ids.peca_input.text = ''

        except Exception as error:
            print(f'Houve um erro - {error}')
            if self.tipo_busca == 'data':
                self.ids.finalizar_alteracao.text = 'Faça outra Busca\n  Para atualizar'
            elif self.tipo_busca == 'index':
                self.buscarPesquisa()
            else:
                self.ids.finalizar_alteracao.text = 'Faça outra Busca\n  Para atualizar'


class LimparTodasAsListas(Screen):
    """
    """
    pass


class ConsultaDeListas(Screen):
    """
    --> Opção para consultar os dados existentes, porém se faz necessário escolher as
    lista que deseja consultar,
    poderá escolher entre: RD Marcas, Perfumaria, Dermo ou todas ao mesmo tempo.
    """
    pass


class ConsultaRDMarcas(Screen):
    """
    --> Opção que consulta a lista RD Marcas.
    """
    pass


class ConsultaPerfumaria(Screen):
    """
    --> Opção que consulta a lista Perfumaria.
    """
    pass


class ConsultaDermo(Screen):
    """
    --> Opção que consulta a lista Dermo.
    """
    pass


class ConsultaTodasAsListas(Screen):
    """
    --> Opção que consulta todas as listas.
    """
    pass


class CriarBackup(Screen):
    """
    Opção para fazer backup dos dados existentes, porém se faz necessário escolher as
    lista que deseja fazer o backup,
    poderá escolher entre: RD Marcas, Perfumaria, Dermo ou todas ao mesmo tempo.
    """

    def __init__(self, **kw):
        super().__init__()
        self.button = None

    def fazerBackup_popup(self, button):
        """
        --> Função que mostra um Popup de confirmação antes de prosseguir com a exclusão da lista.
        """
        try:
            self.button = button

            content = BoxLayout(orientation='vertical', padding=10)
            label = Label(text='Confirma o Backup?')

            close_button = Button(text='Cancelar', size_hint=(None, None), size=(313, 50))
            confirm_button = Button(text='Confirmar', size_hint=(None, None), size=(313, 50))

            content.add_widget(label)
            content.add_widget(close_button)
            content.add_widget(confirm_button)

            popup = Popup(title='Aviso', content=content, size_hint=(None, None), size=(360, 280))

            close_button.bind(on_release=popup.dismiss)

            if self.button == 'RD MARCAS' or self.button == 'PERFUMARIA' or self.button == 'DERMO':
                confirm_button.bind(on_release=lambda btn: self.realizarBackup())
            else:
                confirm_button.bind(on_release=lambda btn: self.realizarBackup_All())

            confirm_button.bind(on_release=popup.dismiss)
            popup.open()

        except Exception as error:
            print(error)
            popupError()

    def realizarBackup(self):
        try:
            # Pega a data formatada no dia atual
            hora = datetime.datetime.now()
            date = datetime.datetime.now()
            date = datetime.datetime.date(date)
            datahoje = date.strftime("%d-%m-%Y")
            horahoje = hora.strftime("%H;%M;%S")

            if self.button == 'RD MARCAS':
                nomeArquivoRD = f"BackupRDMARCAS-{datahoje}-{horahoje}"
                df_lista_RDMarcas = pd.read_excel('storage/listaRDMarcas.xlsx')
                df_lista_RDMarcas = formataLista(df_lista_RDMarcas, self.button)
                df_lista_RDMarcas.to_excel(f'backup/RDMarcas/{nomeArquivoRD}.xlsx', index=False)

            elif self.button == 'PERFUMARIA':
                nomeArquivoPERFUMARIA = f"BackupPERFUMARIA-{datahoje}-{horahoje}"
                df_lista_Perfumaria = pd.read_excel('storage/listaPerfumaria.xlsx')
                df_lista_Perfumaria = formataLista(df_lista_Perfumaria, self.button)
                df_lista_Perfumaria.to_excel(f'backup/Perfumaria/{nomeArquivoPERFUMARIA}.xlsx', index=False)

            elif self.button == 'DERMO':
                nomeArquivoDERMO = f"BackupDERMO-{datahoje}-{horahoje}"
                df_lista_Dermo = pd.read_excel('storage/listaDermo.xlsx')
                df_lista_Dermo = formataLista(df_lista_Dermo, self.button)
                df_lista_Dermo.to_excel(f'backup/Dermo/{nomeArquivoDERMO}.xlsx', index=False)

            popup_Confirmacao_Backup()

        except Exception as error:
            print(error)
            popupError()

    @staticmethod
    def realizarBackup_All():
        try:
            # Pega a data formatada no dia atual
            hora = datetime.datetime.now()
            date = datetime.datetime.now()
            date = datetime.datetime.date(date)
            datahoje = date.strftime("%d-%m-%Y")
            horahoje = hora.strftime("%H;%M;%S")

            nomeArquivoRD = f"BackupRDMARCAS-{datahoje}-{horahoje}"
            nomeArquivoPERFUMARIA = f"BackupPERFUMARIA-{datahoje}-{horahoje}"
            nomeArquivoDERMO = f"BackupDERMO-{datahoje}-{horahoje}"

            df_lista_RDMarcas = pd.read_excel('storage/listaRDMarcas.xlsx')
            df_lista_Perfumaria = pd.read_excel('storage/listaPerfumaria.xlsx')
            df_lista_Dermo = pd.read_excel('storage/listaDermo.xlsx')

            df_lista_RDMarcas = formataLista(df_lista_RDMarcas, button='RD MARCAS')
            df_lista_Perfumaria = formataLista(df_lista_Perfumaria, button='PERFUMARIA')
            df_lista_Dermo = formataLista(df_lista_Dermo, button='DERMO')

            df_lista_RDMarcas.to_excel(f'backup/RDMarcas/{nomeArquivoRD}.xlsx', index=False)
            df_lista_Perfumaria.to_excel(f'backup/Perfumaria/{nomeArquivoPERFUMARIA}.xlsx', index=False)
            df_lista_Dermo.to_excel(f'backup/Dermo/{nomeArquivoDERMO}.xlsx', index=False)
            popup_Confirmacao_Backup()

        except Exception as error:
            print(error)
            popupError()


class FecharPrograma(Screen):
    """
    Opção simples para fechar o programa com segurança sem medo de perder os
    dados ou interromper no meio do processo.
    """
    pass


class Tela(App):

    def build(self):
        if sistema_windows:
            Window.size = (379, 810)
        self.title = 'ConsultaDeMetas_v2.0'
        adm = ScreenManager()
        return adm


if __name__ == '__main__':
    Tela().run()
