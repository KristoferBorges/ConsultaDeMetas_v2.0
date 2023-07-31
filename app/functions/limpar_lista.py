from modulo import popup_Confirmacao_Exclusao
from time import sleep
import platform
import pandas as pd
from modulo import popupError
from kivy.uix.screenmanager import Screen
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.popup import Popup
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.properties import NumericProperty
from openpyxl import load_workbook

# Verifica se o usuário está usando Windows
if platform.system() == "Windows":
    sistema_windows = True
    font_column = 18
    font_row = 16
    font_button = 35
    font_text = 35
    font_text_menu = 48
    font_title = 60

else:
    sistema_windows = False
    font_column = 20
    font_row = 19
    font_button = 55
    font_text = 40
    font_text_menu = 60
    font_title = 80


class LimparDados(Screen):
    """
    Opção para limpar dos dados, porém se faz necessário escolher as
    lista que deseja limpar os dados,
    poderá escolher entre: RD Marcas, Perfumaria, Dermo ou todas ao mesmo tempo.
    """

    font_column = NumericProperty(font_column)
    font_row = NumericProperty(font_row)
    font_button = NumericProperty(font_button)
    font_text = NumericProperty(font_text)
    font_text_menu = NumericProperty(font_text_menu)
    font_title = NumericProperty(font_title)

    def apagarLista_popup(self):
        """
        --> Função que mostra um Popup de confirmação antes de prosseguir com a exclusão das listas.
        """

        content = BoxLayout(orientation='vertical', padding=10)
        label = Label(text='Confirma a exclusão das Listas?')

        close_button = Button(text='Cancelar', size_hint=(1.0, 0.2))
        confirm_button = Button(text='Confirmar', size_hint=(1.0, 0.2))

        content.add_widget(label)
        content.add_widget(close_button)
        content.add_widget(confirm_button)

        popup = Popup(title='Aviso', content=content, size_hint=(0.7, 0.5))

        close_button.bind(on_release=popup.dismiss)

        confirm_button.bind(on_release=lambda btn: self.apagarLista())
        confirm_button.bind(on_release=popup.dismiss)
        popup.open()

    def apagarLista(self):
        """
        --> Função que apaga todas as listas após confirmar no popup o procedimento.
        """

        try:
            # Exclusão RD MARCAS
            # Carrega o arquivo
            lista = 'storage/listaRDMarcas.xlsx'
            calculo = 'storage/lista_calc_RDMarcas.xlsx'
            bk_lista = load_workbook(lista)
            bk_calculo = load_workbook(calculo)

            # Pega a primeira planilha do arquivo de lista
            sheet_lista = bk_lista.active

            # Pega a primeira planilha do arquivo de cálculo
            sheet_calculo = bk_calculo.active

            # Exclui as linhas tirando a primeira (Nome das colunas/Key)
            sheet_lista.delete_rows(2, sheet_lista.max_row)
            sheet_calculo.delete_rows(2, sheet_calculo.max_row)

            # Salva as alterações
            bk_lista.save(lista)
            bk_calculo.save(calculo)

            # Passa as alterações para uma variável
            df_lista_RDMarcas = pd.read_excel(lista)
            calc_lista_RDMarcas = pd.read_excel(calculo)

            # Salva o arquivo com as alterações no DataFrame
            df_lista_RDMarcas.to_excel('storage/listaRDMarcas.xlsx', index=False)
            calc_lista_RDMarcas.to_excel('storage/lista_calc_RDMarcas.xlsx', index=False)

            # Exclusão PERFUMARIA
            # Carrega o arquivo
            lista = 'storage/listaPerfumaria.xlsx'
            calculo = 'storage/lista_calc_Perfumaria.xlsx'
            bk_lista = load_workbook(lista)
            bk_calculo = load_workbook(calculo)

            # Pega a primeira planilha do arquivo de lista
            sheet_lista = bk_lista.active

            # Pega a primeira planilha do arquivo de cálculo
            sheet_calculo = bk_calculo.active

            # Exclui as linhas tirando a primeira (Nome das colunas/Key)
            sheet_lista.delete_rows(2, sheet_lista.max_row)
            sheet_calculo.delete_rows(2, sheet_calculo.max_row)

            # Salva as alterações
            bk_lista.save(lista)
            bk_calculo.save(calculo)

            # Passa as alterações para uma variável
            df_lista_Perfumaria = pd.read_excel(lista)
            calc_lista_Perfumaria = pd.read_excel(calculo)

            # Salva o arquivo com as alterações no DataFrame
            df_lista_Perfumaria.to_excel('storage/listaPerfumaria.xlsx', index=False)
            calc_lista_Perfumaria.to_excel('storage/lista_calc_Perfumaria.xlsx', index=False)

            # Exclusão Dermo
            # Carrega o arquivo
            lista = 'storage/listaDermo.xlsx'
            calculo = 'storage/lista_calc_Dermo.xlsx'
            bk_lista = load_workbook(lista)
            bk_calculo = load_workbook(calculo)

            # Pega a primeira planilha do arquivo de lista
            sheet_lista = bk_lista.active

            # Pega a primeira planilha do arquivo de cálculo
            sheet_calculo = bk_calculo.active

            # Exclui as linhas tirando a primeira (Nome das colunas/Key)
            sheet_lista.delete_rows(2, sheet_lista.max_row)
            sheet_calculo.delete_rows(2, sheet_calculo.max_row)

            # Salva as alterações
            bk_lista.save(lista)
            bk_calculo.save(calculo)

            # Passa as alterações para uma variável
            df_lista_Dermo = pd.read_excel(lista)
            calc_lista_Dermo = pd.read_excel(calculo)

            # Salva o arquivo com as alterações no DataFrame
            df_lista_Dermo.to_excel('storage/listaDermo.xlsx', index=False)
            calc_lista_Dermo.to_excel('storage/lista_calc_Dermo.xlsx', index=False)

            # POPUP DE FINALIZAÇÃO
            sleep(0.3)
            popup_Confirmacao_Exclusao()

        except Exception as error:
            print(error)
            popupError()


class LimparTodasAsListas(Screen):
    """
    """
    pass
